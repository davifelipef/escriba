import os
import sys
if sys.__stdout__ is None or sys.__stderr__ is None:
    os.environ['KIVY_NO_CONSOLELOG'] = '1'
import datetime
import openpyxl
import pygame
import locale
import threading
locale.setlocale(locale.LC_TIME, 'pt_BR.utf8')
from kivy.core.window import Window
from kivymd.app import MDApp
from kivy.lang import Builder
from kivy.uix.boxlayout import BoxLayout
from kivymd.uix.snackbar import Snackbar
from collections import defaultdict
from openpyxl.utils import get_column_letter
from kivymd.uix.pickers import MDDatePicker
from kivy.clock import Clock

# loads the layout file
Builder.load_file('graphics/layout.kv')

# main class of the program
class MainScreen(BoxLayout):

    # Starts important variables
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        # setting up DATE and TIME variables to automatically update
        self.DATE = datetime.datetime.now().strftime("%d/%m/%Y")
        self.TIME = datetime.datetime.now().strftime("%H:%M:%S")
        # initiating the variables used by the snackbar that counts clicks
        self.snackbar = None
        self.counter = 0
        self.last_button_press_time = None

    # Plays a click sound
    def click_sound(self):
        try:
            pygame.mixer.init()
            sound_file = os.path.join(os.getcwd(), 'sounds/click.wav')
            sound = pygame.mixer.Sound(sound_file)
            sound.play()
        except FileNotFoundError:
            sound = None

    # Plays an alert sound
    def alert_sound(self):
        try:
            pygame.mixer.init()
            sound_file = os.path.join(os.getcwd(), 'sounds/alert.wav')
            sound = pygame.mixer.Sound(sound_file)
            sound.set_volume(0.3)
            sound.play()
        except FileNotFoundError:
            sound = None

    # Plays a sound that represents data being erased
    def eraser_sound(self):
        try:
            pygame.mixer.init()
            sound_file = os.path.join(os.getcwd(), 'sounds/eraser.wav')
            sound = pygame.mixer.Sound(sound_file)
            sound.set_volume(0.3)
            sound.play()
        except FileNotFoundError:
            sound = None

    # Plays a confirmation sound
    def data_saved_sound(self):
        try:
            pygame.mixer.init()
            sound_file = os.path.join(os.getcwd(), 'sounds/data_saved.wav')
            sound = pygame.mixer.Sound(sound_file)
            sound.play()
        except FileNotFoundError:
            sound = None

    # Opens the assets folder for easy access to the excel file
    def open_assets_folder(self, *args):
        # Check if the script is running as an exe file or as a python script
        if getattr(sys, 'frozen', False):
            # If running as an exe file, get the directory where the exe file is located
            script_dir = os.path.dirname(sys.executable)
        else:
            # If running as a python script, get the current working directory
            script_dir = os.getcwd()
        
        # sets the path to the file
        file_path = "assets/data.xlsx"
        
        # Join the script directory with the folder name
        full_file_path = os.path.join(script_dir, file_path)
        
        # Open the folder in the default file explorer of the operating system
        os.startfile(full_file_path)

    # Creates a new thread to run the report generation on the background
    def generate_report_thread(self):
        # starts the report creation thread
        threading.Thread(target=self.generate_report).start()

    # Generate report method
    def generate_report(self, instance=None):
        def is_time_in_period(time_value, period):
            if period == "Manhã":
                return "08:00:00" <= time_value <= "11:59:59"
            elif period == "Tarde":
                return "12:00:00" <= time_value <= "17:59:59"
            elif period == "Noite":
                return "18:00:00" <= time_value <= "20:59:59"
            else:
                return False

        input_file = "assets/data.xlsx"
        workbook = openpyxl.load_workbook(input_file)
        sheet = workbook.active

        public_types = ["CEI", "EMEI", "EMEF", "ETEC", "Comunidade", "Funcionário"]
        ages = ["Até 12", "13 a 17", "18 a 59", "60 ou mais"]
        time_of_day = ["Manhã", "Tarde", "Noite"]

        public_type_counts = defaultdict(int)
        age_counts = defaultdict(int)
        time_counts = defaultdict(int)

        dates = set()
        data = []

        for row in sheet.iter_rows(values_only=True, min_row=2):
            date_str, time_str, age, public_type = row[:4]

            if len(date_str.split('/')) == 3:
                dates.add(date_str)
                data.append((date_str, time_str, age, public_type))

                if public_type in public_types:
                    public_type_counts[(date_str, public_type)] += 1

                if age in ages:
                    age_counts[(date_str, age)] += 1

                for period in time_of_day:
                    if is_time_in_period(time_str, period):
                        time_counts[(date_str, period)] += 1

        valid_dates = sorted(dates, key=lambda date: tuple(map(int, date.split('/'))))

        desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
        output_file = f"consolidado-{self.DATE.replace('/', '-')}.xlsx"
        output_workbook = openpyxl.Workbook()
        output_sheet = output_workbook.active

        # Set the dates as the column headers
        for index, date in enumerate(valid_dates):
            column_letter = get_column_letter(index + 2)
            output_sheet[column_letter + "1"] = date

         # Sets the public types as the first set of rows
        public_types = ["CEI", "EMEI", "EMEF", "ETEC", "Comunidade", "Funcionário"]
        for index, public_type in enumerate(public_types):
            # Sets the cell A2 as the start of the public type rows so A7 holds the last row
            output_sheet["A" + str(index + 2)] = public_type

        # Sets the cell A8 as the total row to sum all the public types numbers
        output_sheet["A8"] = "Total"

        # Sets the ages as the second set of rows
        ages = ["Até 12", "13 a 17", "18 a 59", "60 ou mais"]
        for index, age in enumerate(ages):
            # Sets the cell A9 as the start of the age rows so A12 holds the last row
            output_sheet["A" + str(index + 9)] = age

        # Sets the cell A13 as the total row to sum all the age numbers
        output_sheet["A13"] = "Total"

        # Sets the period of the day as the third and last set of rows
        time_of_day = ["Manhã", "Tarde", "Noite"]
        for index, time in enumerate(time_of_day):
            # Sets the cell A14 as the start of the age rows so A16 holds the last row
            output_sheet["A" + str(index + 14)] = time

        # Sets the cell A17 as the total row to sum all the period of the day numbers
        output_sheet["A17"] = "Total"

        # Populate counts for public types
        for index, date in enumerate(valid_dates):
            column_letter = get_column_letter(index + 2)
            for i, public_type in enumerate(public_types):
                output_sheet[column_letter + str(i + 2)] = public_type_counts.get((date, public_type), 0)

            output_sheet[column_letter + "8"] = f"=SUM({column_letter}2:{column_letter}7)"

        # Populate counts for ages
        for index, date in enumerate(valid_dates):
            column_letter = get_column_letter(index + 2)
            for i, age in enumerate(ages):
                output_sheet[column_letter + str(i + 9)] = age_counts.get((date, age), 0)

            output_sheet[column_letter + "13"] = f"=SUM({column_letter}9:{column_letter}12)"

        # Populate counts for time of day
        for index, date in enumerate(valid_dates):
            column_letter = get_column_letter(index + 2)
            for i, time in enumerate(time_of_day):
                output_sheet[column_letter + str(i + 14)] = time_counts.get((date, time), 0)

            output_sheet[column_letter + "17"] = f"=SUM({column_letter}14:{column_letter}16)"

        output_workbook.save(os.path.join(desktop_path, output_file))
        # calls the snackbar that informs the report is ready
        Clock.schedule_once(self.report_ready_snackbar)  # Note: Define or implement your 'report_ready_snackbar' function

    # Snackbar that informs the report is ready
    def report_ready_snackbar(self, dt):
        self.snackbar = Snackbar(
                            text="O relatório está pronto!", 
                            bg_color=(0, 0.5, 0, 1), # green
                            font_size="16sp"
                        )
        self.snackbar.open()
        Clock.schedule_once(self.dismiss_snackbar, 8)
        
    # Erases all the selected buttons and reset their colors
    def eraser(self):
        # make its color blue
        self.ids.auto_button.md_bg_color = 'blue'
        # makes the manual button color grey
        self.ids.manual_button.md_bg_color = (0.35, 0.35, 0.35, 1)
        # makes all the period of the day button grey
        self.ids.morning_button.md_bg_color = (0.35, 0.35, 0.35, 1)
        self.ids.afternoon_button.md_bg_color = (0.35, 0.35, 0.35, 1)
        self.ids.night_button.md_bg_color = (0.35, 0.35, 0.35, 1)
        # keeps all the manual date and period of the day buttons disabled
        self.ids.select_date.disabled = True
        self.ids.calendar_button.disabled = True
        self.ids.select_day_period.disabled = True
        self.ids.morning_button.disabled = True
        self.ids.afternoon_button.disabled = True
        self.ids.night_button.disabled = True
        # makes all the age selection buttons grey again
        self.ids.twelve_yo_button.md_bg_color = (0.35, 0.35, 0.35, 1) 
        self.ids.thirteen_til_seventeen_yo_button.md_bg_color = (0.35, 0.35, 0.35, 1)
        self.ids.eighteen_til_fiftynine_yo_button.md_bg_color = (0.35, 0.35, 0.35, 1) 
        self.ids.sixty_yo_button.md_bg_color = (0.35, 0.35, 0.35, 1)
        # makes all the public selection buttons grey again
        self.ids.cei_button.md_bg_color = (0.35, 0.35, 0.35, 1) 
        self.ids.emei_button.md_bg_color = (0.35, 0.35, 0.35, 1) 
        self.ids.emef_button.md_bg_color = (0.35, 0.35, 0.35, 1)
        self.ids.etec_button.md_bg_color = (0.35, 0.35, 0.35, 1)
        self.ids.community_button.md_bg_color = (0.35, 0.35, 0.35, 1) 
        self.ids.employee_button.md_bg_color = (0.35, 0.35, 0.35, 1)
        # plays an erasing sound
        self.eraser_sound()

    # Handles what button was clicked between auto and manual
    def toggle_date_and_time(self, button):
        # declares variables associated with the two possible buttons
        auto_button = self.ids.auto_button
        manual_button = self.ids.manual_button
        # if the auto button is clicked
        if button == auto_button:
            # returns the date to the current date
            self.DATE = datetime.datetime.now().strftime("%d/%m/%Y")
            print(self.DATE)
            # returns the time to the current time
            self.TIME = datetime.datetime.now().strftime("%H:%M:%S")
            print(self.TIME)
            # info message
            snackbar = Snackbar(
                            text=f"Data alterada para {self.DATE}.", 
                            duration=4.0,
                            bg_color=(0, 0.5, 0, 1), # green
                            font_size="16sp"
                        )
            snackbar.open()
            # make its color blue
            button.md_bg_color = 'blue'
            # makes the manual button color grey
            manual_button.md_bg_color = (0.35, 0.35, 0.35, 1)
            # makes all the period of the day button grey
            self.ids.morning_button.md_bg_color = (0.35, 0.35, 0.35, 1)
            self.ids.afternoon_button.md_bg_color = (0.35, 0.35, 0.35, 1)
            self.ids.night_button.md_bg_color = (0.35, 0.35, 0.35, 1)
            # keeps all the manual date and period of the day buttons disabled
            self.ids.select_date.disabled = True
            self.ids.calendar_button.disabled = True
            self.ids.select_day_period.disabled = True
            self.ids.morning_button.disabled = True
            self.ids.afternoon_button.disabled = True
            self.ids.night_button.disabled = True
        # if the manual button is clicked    
        elif button == manual_button:
            # make its color blue
            button.md_bg_color = 'blue'
            # makes the auto button color grey
            auto_button.md_bg_color = (0.35, 0.35, 0.35, 1)
            # enable sall the manual date and period of the day buttons
            self.ids.select_date.disabled = False
            self.ids.calendar_button.disabled = False
            self.ids.select_day_period.disabled = False
            self.ids.morning_button.disabled = False
            self.ids.afternoon_button.disabled = False
            self.ids.night_button.disabled = False

    # Handles the time of the day buttons
    def on_manual_period(self, instance):
        buttons = [self.ids.morning_button, self.ids.afternoon_button, self.ids.night_button]
        # toggle the selected button's color between blue and grey
        if instance.md_bg_color == [0, 0, 1, 1]:
            instance.md_bg_color = [0.35, 0.35, 0.35, 1]
        else:
            for button in buttons:
                button.md_bg_color = [0.35, 0.35, 0.35, 1]
            instance.md_bg_color = [0, 0, 1, 1]
            self.TIME = instance.text 

    # Handles the on save used in the date picker function
    def on_save_date(self, instance, value, date_range):
        self.DATE = value.strftime("%d/%m/%Y")
        # play an alert sound
        self.alert_sound()
        snackbar = Snackbar(
                            text=f"Data alterada para {self.DATE}.", 
                            duration=4.0,
                            bg_color=(0, 0.5, 0, 1), # green
                            font_size="16sp"
                        )
        snackbar.open()
        
    # Opens the date picker widget
    def show_date_picker(self):
        date_dialog = MDDatePicker()
        date_dialog.bind(on_save=self.on_save_date)
        date_dialog.open()

    # Handles the age selection buttons
    def age_selection(self, instance):
        buttons = [self.ids.twelve_yo_button, self.ids.thirteen_til_seventeen_yo_button,
                self.ids.eighteen_til_fiftynine_yo_button, self.ids.sixty_yo_button]
        # toggle the selected button's color between blue and grey
        if instance.md_bg_color == [0, 0, 1, 1]:
            instance.md_bg_color = [0.35, 0.35, 0.35, 1]
        else:
            for button in buttons:
                button.md_bg_color = [0.35, 0.35, 0.35, 1]
            instance.md_bg_color = [0, 0, 1, 1]

    # Handles the public selection buttons
    def public_selection(self, instance):
        buttons = [self.ids.cei_button, self.ids.emei_button, self.ids.emef_button,
                   self.ids.etec_button, self.ids.community_button, self.ids.employee_button]
        # toggle the selected button's color between blue and grey
        if instance.md_bg_color == [0, 0, 1, 1]:
            instance.md_bg_color = [0.35, 0.35, 0.35, 1]
        else:
            for button in buttons:
                button.md_bg_color = [0.35, 0.35, 0.35, 1]
            instance.md_bg_color = [0, 0, 1, 1]

    # Creates a new thread to run the save data on the background
    def save_data_thread(self):
        threading.Thread(target=self.save_data).start()

    # Save data method
    def save_data(self, instance=None, *args):
        # checks for the data where the file will be saved
        wb_path = 'assets/data.xlsx'
        # if the file isn't there, creates the file and sets up the sheet headers
        if not os.path.exists(wb_path):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Data"
            ws['A1'] = "Data"
            ws['B1'] = "Hora"
            ws['C1'] = "Faixa Etária"
            ws['D1'] = "Tipo de Público"
            wb.save(wb_path)
        # if the file is already there, loads it as is
        else:
            wb = openpyxl.load_workbook(wb_path)
            ws = wb.active

        # gets the two sets of row that hold the information about the public age and type
        button_sets = [self.ids.public_age_row, self.ids.public_type_row]
        # list that will hold the selected button options
        selected_options = []
        for button_set in button_sets:
            for button_id in button_set.children:
                # check if the button color is blue (which means that it's selected)
                if button_id.md_bg_color == [0, 0, 1, 1]:
                    # if the button color is blue, add its text to the list
                    selected_options.append(button_id.text)
        
            # checks if the TIME variable was set up manually
            if self.TIME == "Manhã":
                pass
            elif self.TIME == "Tarde":
                pass
            elif self.TIME == "Noite":
                pass
            # if the self.TIME variable was not set manually, gets the current time the save button was clicked
            else:
                self.TIME = datetime.datetime.now().strftime("%H:%M:%S")

        # sets up the row order in the excel file
        row = [None] * 5
        # writes the date to the first column
        row[0] = self.DATE
        # writes the time to the second column
        row[1] = self.TIME
        for i in range(min(2, len(selected_options))):
            if i == 0:
                # writes the age to the third column
                row[2] = selected_options[i]
            elif i == 1:
                # writes the type of public to the fourth column
                row[3] = selected_options[i]

        # checks if there is an age and a public information currently selected
        if selected_options and len(selected_options) >= 2:

            # saves the information to the excel file
            ws.append(row)
            wb.save(wb_path)

            # plays a sound when the data is saved
            self.data_saved_sound()

            # calls the info snackbar
            Clock.schedule_once(self.save_successful_snackbar)
            
        else:
            # calls an alert snackbar
            Clock.schedule_once(self.save_alert_snackbar)

    # Snackbar that informs the data was saved successfully
    def save_successful_snackbar(self, dt):
        ''' snackbar that shows that the data was saved and sums the save count if 
        it is saved again in less than 3sec'''
        if self.snackbar is None:
            self.counter = 1
            text = f"Dados salvos x{self.counter}"
            self.snackbar = Snackbar(
                            text=text, 
                            bg_color=(0, 0.5, 0, 1), # green
                            font_size="16sp"
                        )
            self.snackbar.open()
            Clock.schedule_once(self.dismiss_snackbar, 3)
        else:
            self.counter += 1
            text = f"Dados salvos x{self.counter}"
            self.snackbar.text = text
            if self.snackbar.duration is None:
                remaining_time = 3
            else:
                remaining_time = self.snackbar.duration - self.last_button_press_time + Clock.get_time()
            if remaining_time < 3:
                self.snackbar.duration += 3
            else:
                self.snackbar.duration = remaining_time + 3
            Clock.unschedule(self.dismiss_snackbar)
            Clock.schedule_once(self.dismiss_snackbar, self.snackbar.duration - remaining_time)

        self.last_button_press_time = Clock.get_time()

    # Snackbar that informs the required selections are missing
    def save_alert_snackbar(self, dt):
        # plays a sound when the alert message is displayed
        self.alert_sound()
        alert_snackbar = Snackbar(
                            text='Selecione a faixa etária e o público!', 
                            bg_color=(1, 0, 0, 1), # red
                            font_size="16sp"
                        )
        alert_snackbar.open()

    # Dismisses the snackbar
    def dismiss_snackbar(self, dt):
        if self.snackbar:
            self.snackbar.dismiss()
            self.snackbar = None
            self.counter = 0

# app class
class Scribe(MDApp):
 
    def build(self):
        # sets the app window title
        self.title = 'Escriba - Versão 0.4 - 10/11/2023'
        # set the taskbar icon
        Window.set_icon('graphics/app.ico')
        # set the window icon
        self.icon = 'graphics/app.ico'
        #exibts the layout above in the user screen
        return MainScreen()      

# runs the program
if __name__ == '__main__':
    Scribe().run()
